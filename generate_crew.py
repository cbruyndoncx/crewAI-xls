import sys
import os
from textwrap import dedent

import openpyxl
from jinja2 import Environment, FileSystemLoader

models = {}
models['temperature']=0.1

import re
def snake_case(s):
    # empty string do nothing
    if s is None:
        return
    
    # Remove all non-word characters (everything except numbers and letters)
    s = re.sub(r"[^\w\s]", '', s)
    
    # Replace all runs of whitespace with a single underscore
    s = re.sub(r"\s+", '_', s)
    
    # Convert to lowercase
    return s.lower()

def create_directory(directory_path):
    # Check if the directory already exists
    if not os.path.exists(directory_path):
        # If it doesn't exist, create it (including any intermediate directories)
        os.makedirs(directory_path)
        print(f"Directory '{directory_path}' was created.")
    else:
        print(f"Directory '{directory_path}' already exists.")

def replace_none_with_empty_string(records):
    for record in records:
        for key in record:
            if record[key] is None:
                record[key] = ''
    return records

def read_variables_sheet(sheet):
    
    # Read variables from Excel file into a dictionary
    variables = {}

    for i, cell in enumerate(sheet[1], start=1):
        key = cell.value  # Assuming headers are keys for your template
        value = sheet.cell(row=2, column=i).value  # Read value from second row
        variables[key] = value
    
    # Read headers from the first row to use as keys for each dictionary
    headers = [cell.value for cell in sheet[1]]

    # Initialize an empty list to store all records as dictionaries
    all_records = []

    # Loop over all rows starting from the second row (since first row is header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Create a dictionary for each record (row)
        record = {headers[i]: cell for i, cell in enumerate(row)}
        # Add the dictionary to the list of all records
        all_records.append(record)

    return all_records

def read_variables_xls(template_filename, select_crew, select_job, crews_dir):

    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(template_filename)
    
    # Set up the Jinja2 environment and specify the directory containing templates
    env = Environment(loader=FileSystemLoader('templates'))
        #sheet = workbook[sheet_name]
    crew_agent_list_template = ''
    crew_task_list_template = ''
    
    agent_records = read_variables_sheet(workbook['agents'])
    task_records = read_variables_sheet(workbook['tasks'])
    custom_records = read_variables_sheet(workbook['custom'])
    crew_records = read_variables_sheet(workbook['crews'])
    crew_member_records = read_variables_sheet(workbook['crewmembers'])
    # Now `all_records` is a list of dictionaries with details for all records in the different sheets
    
    # TASKS 
    # Open the file in write mode ('w'). If the file doesn't exist, it will be created.
    # If you want to append to an existing file instead, use 'a' mode.
    with open(f"{crews_dir}tasks.py", 'w') as file:
        file.write(env.get_template('tasks_class_template.py').render(models))
    
    # Add each task details
    job_records = []
    with open(f"{crews_dir}tasks.py", 'a') as file:
        task_records=replace_none_with_empty_string(task_records)
        for record in task_records:
            record['job']=snake_case(record['job'])
            if record['job'] == select_job:
                # Process the record as needed
                record['task_name'] = snake_case(record['task'])
                record['context'] = snake_case(record['context'])
                record['crews_dir'] = crews_dir
                file.write(env.get_template('task_template.py').render(record))
                if record['assigned_agent'] == '' :
                    record['assigned_agent_name'] = 'None'
                else:
                    record['assigned_agent_name'] = snake_case(record['assigned_agent'])

                crew_task_list_template+=env.get_template('crew_task_list_template.py').render(record)
                job_records.append(record)
    # CUSTOM VARS 
    # Open the file in write mode ('w'). If the file doesn't exist, it will be created.
    # If you want to append to an existing file instead, use 'a' mode.
    # ToDo

    # CREW MEMBERS
    # Open the file in write mode ('w'). If the file doesn't exist, it will be created.
    # If you want to append to an existing file instead, use 'a' mode.
    member_records = []

    def add_member_record(select_crew, assigned_agent_name):
        # Create a new crew info record
        new_record = {
            'crew': snake_case(select_crew),
            'crewmember': snake_case(assigned_agent_name)
        }

        # Check if this crew info already exists in member records
        if not any(rec['crew'] == new_record['crew'] and rec['crewmember'] == new_record['crewmember'] for rec in member_records):
            # If it doesn't exist, append it to member_records
            member_records.append(new_record)

    for record in crew_member_records:
        if record['crew'] == select_crew:
            add_member_record(select_crew, record['crewmember'])

    for record in job_records:
        if record['assigned_agent'] != '':
            add_member_record(select_crew, record['assigned_agent_name'])
    #print(member_records)

    # AGENTS
    # Open the file in write mode ('w'). If the file doesn't exist, it will be created.
    # If you want to append to an existing file instead, use 'a' mode.
    with open(f"{crews_dir}agents.py", 'w') as file:
        file.write(env.get_template('agents_class_template.py').render(models))

    # Add each agent details
    with open(f"{crews_dir}agents.py", 'a') as file:
        agent_records=replace_none_with_empty_string(agent_records)
        for record in agent_records:
            for member in member_records:
                record['agent_name'] = snake_case(record['agent'])
                if member['crewmember'] == record['agent_name']: 
                    record['crews_dir'] = crews_dir
                    file.write(env.get_template('agent_template.py').render(record))
                    crew_agent_list_template+=env.get_template('crew_agent_list_template.py').render(record)

    # CREW
    # Open the file in write mode ('w'). If the file doesn't exist, it will be created.
    # If you want to append to an existing file instead, use 'a' mode.
    crew_agent_list = ','.join([str(snake_case(record['crewmember'])) for record in member_records])
    crew_task_list = ','.join([str(snake_case(record['task'])) for record in job_records])

    for record in crew_records:
        record['crew_name'] = snake_case(record['crew'])
        if record['crew_name'] == select_crew:
            record['crew_agent_list_template']=crew_agent_list_template
            record['crew_task_list_template']=crew_task_list_template
            record['crew_agent_list']=crew_agent_list
            record['crew_task_list']=crew_task_list

            result=env.get_template('crew_class_template.py').render(record)

            with open(f"{crews_dir}crew.py", 'w') as file:
                file.write(result)
            return
    print(f"crew {select_crew} is not defined")
    sys.exit(1)  # Exit with a non-zero value to indicate failure


# This is the main function that you will use to run your custom crew.
if __name__ == "__main__":
    print("## Welcome to Crew AI XLS Crew ##")
    print("---------------------------------")
    input_crew = snake_case(input(dedent("""Specify your crew """)))
    input_job = snake_case(input(dedent("""Specify your job: """)))

    crews_dir = f"./crews/{input_crew}-{input_job}/"
    create_directory(crews_dir)

    read_variables_xls('vars_in.xlsx',input_crew, input_job, crews_dir)    

    # If everything is successful, you can either omit sys.exit() or explicitly exit with zero
    sys.exit(0)  # This line is optional since Python exits with 0 by default
