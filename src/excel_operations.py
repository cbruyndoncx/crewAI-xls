# excel_operations.py

import os
import re
import glob
from datetime import datetime
import openpyxl
import logging

# Function to sanitize a string for Excel
def sanitize_for_excel(value):
    if not isinstance(value, str):
        return value
    
    # Remove leading and trailing whitespaces
    sanitized_value = value.strip()
    
    # Escape characters that are interpreted by Excel as formulas
    if sanitized_value.startswith(('=', '+', '-', '@')):
        sanitized_value = "'" + sanitized_value
    
    # Remove non-printable characters
    sanitized_value = re.sub(r'[^\x20-\x7E]+', '', sanitized_value)
    sanitized_value = sanitized_value.replace('"', '').replace("'", '')
    
    return sanitized_value

def open_workbook(xls_file_path):
    """
    Create a new workbook or load an existing one.
    """
    try:
        wb = openpyxl.load_workbook(xls_file_path)
        
    except FileNotFoundError:
        logging.info(f"The file {xls_file_path} does not exist. Creating a new workbook.")
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet()

    return wb

def save_workbook(wb, xls_file_path):

    wb.save(xls_file_path)

    return f"Workbook saved as {xls_file_path}"

def write_log_sheet(filename, input, output, final, metrics):

    # Create a new workbook or load an existing one
    wb = open_workbook(filename)

    # Select the default sheet
    log_sheet = wb.active

    # Add column headers
    log_sheet.append(["Timestamp", "Input","Output", "Final Result", "Metrics"])

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
    # Append a row
    log_sheet.append([timestamp, input, sanitize_for_excel(output), sanitize_for_excel(final),sanitize_for_excel(metrics)])

    # Save the workbook to an xlsx file
    logging.info(save_workbook(wb, filename))

def add_md_files_to_log_sheet(filename, directory):
    
    # Create a new workbook or load an existing one
    wb = open_workbook(filename)

    # Find all markdown files with the .md extension
    # Check for permission and existence of directory
    if os.path.exists(directory):
        md_files = glob.glob(os.path.join(directory, '*.md'))
        logging.info(md_files)
    else:
        logging.warning("Directory does not exist or cannot be accessed.")

    #md_files = glob.glob(directory+'/*.md')
    logging.info(md_files)

    # Add content of each markdown file to a new sheet
    for md_file in md_files:
        # Read the content of the current markdown file
        with open(md_file, 'r', encoding='utf-8') as file:
            content = file.read()

        # Create a new sheet with the name of the markdown file (without extension)
        sheet_title = os.path.basename(md_file)[:-3] # Remove the last 3 characters (.md)
        sheet = wb.create_sheet(title=sheet_title)

        # Split content into lines and write each line into a new row
        for row_index, line in enumerate(content.splitlines(), start=1):
            # Assuming the content of each line fits within Excel's cell limit
            line = sanitize_for_excel(line)
            sheet.cell(row=row_index, column=1).value = "'"+line

    logging.info(save_workbook(wb, filename))


def list_xls_files_in_dir(directory):
    """
    This is used to generate the list of template files to choose from
    """
    if directory:
        files_in_directory = os.listdir(directory)
        xlsfiles = [directory + file for file in files_in_directory if file.endswith('.xls') or file.endswith('.xlsx')]
        return xlsfiles
    else:
        return []

def md_list(items):
    return "\n".join(f"* {item}" for item in items)
    
# Function to get distinct values from a named column in a named sheet
def get_distinct_column_values_by_name(workbook_path, sheet_name, column_name):
    """
    Extracts distinct values from a specified named column in an Excel sheet.

    :param workbook_path: The path to the Excel workbook.
    :param sheet_name: The name of the worksheet to use.
    :param column_name: The name of the column to extract values from.
    :return: A set of distinct values.
    """
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[sheet_name]

    # Find the index of the specified column by name
    column_index = 0
    for col in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
        column_index = column_index +1
        if col[0] == column_name:
            break
    else:
        # If no matching header is found, raise an error
        raise ValueError(f"Column with header '{column_name}' not found.")

    # Use a set to store unique values as sets do not allow duplicates
    distinct_values = set()

    # Iterate over all rows in the specified named column (excluding the header)
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        if cell_value is not None:  # Exclude empty cells
            distinct_values.add(cell_value)

    # Close the workbook after processing
    workbook.close()

    return sorted(list(distinct_values))

