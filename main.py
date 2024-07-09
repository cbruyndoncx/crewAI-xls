# main.py

import gradio as gr
import shutil
import os
import sys
import re
import json
from datetime import datetime
import openpyxl
import glob
import pandas as pd

import argparse
from logger import ComplexLogger

from generate_crew import read_variables_xls, snake_case
from importlib import import_module

#######################################
# initialisations
#######################################
crews_dir = ""
CREWS_FOLDER_NAME = "crews"
CREWS_FOLDER = "./" + CREWS_FOLDER_NAME + "/"
XLS_FOLDER = "./xls/"    
OUT_FOLDER = "./out/"
LOG_FOLDER = "./log/"
logfile = LOG_FOLDER + "output.log"
output_log_sheet = OUT_FOLDER + "output_log.xlsx"

#######################################
# Functions
#######################################
def create_default_dir(folder):
    if not os.path.exists(folder):    
        os.mkdir(folder)   

def init_default_dirs():
    create_default_dir(CREWS_FOLDER)  
    create_default_dir(XLS_FOLDER)
    create_default_dir(OUT_FOLDER)
    create_default_dir(LOG_FOLDER)

def init_logging():
    logger = ComplexLogger(logfile)
    logger.reset_logs()
    sys.stdout = ComplexLogger(logfile)
    return logger

def read_logs():
    sys.stdout.flush()
    with open(logfile, "r") as f:
        return f.read()

# Function to sanitize a string for Excel
def sanitize_for_excel(value):
    if not isinstance(value, str):
        return value
    
    # Remove leading and trailing whitespaces
    sanitized_value = value.strip()
    
    # Escape characters that are interpreted by Excel as formulas
    if sanitized_value.startswith(('=', '+', '-', '@')):
        sanitized_value = "'" + sanitized_value
    
    # Remove non-printable characters
    sanitized_value = re.sub(r'[^\x20-\x7E]+', '', sanitized_value)
    sanitized_value = sanitized_value.replace('"', '').replace("'", '')
    
    return sanitized_value

def open_workbook(filename):
    # Create a new workbook or load an existing one
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"The file {filename} does not exist. Creating a new workbook.")
        wb = openpyxl.Workbook(filename)
        sheet = wb.create_sheet()

    return wb

def save_workbook(wb, filename):
        # Save the workbook
    wb.save(filename)

    return f"Workbook saved as {filename}"

def write_log_sheet(filename, input, output, final, metrics):

    # Create a new workbook or load an existing one
    wb = open_workbook(filename)

    # Select the default sheet
    log_sheet = wb.active

    # Add some column headers
    log_sheet.append(["Timestamp", "Input","Output", "Final Result", "Metrics"])

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
    # Append a row
    log_sheet.append([timestamp, input, sanitize_for_excel(output), sanitize_for_excel(final),sanitize_for_excel(metrics)])

    # Save the workbook to an xlsx file
    print(save_workbook(wb, filename))

def add_md_files_to_log_sheet(filename, directory):
    
    # Create a new workbook or load an existing one
    wb = open_workbook(filename)

    # Find all markdown files with the .md extension
    # Check for permission and existence of directory
    if os.path.exists(directory):
        md_files = glob.glob(os.path.join(directory, '*.md'))

        # If you want to include subdirectories
        # md_files = glob.glob(os.path.join(directory, '**/*.md'), recursive=True)

        print(md_files)
    else:
        print("Directory does not exist or cannot be accessed.")

    #md_files = glob.glob(directory+'/*.md')
    print(md_files)

    # Add content of each markdown file to a new sheet
    for md_file in md_files:
        # Read the content of the current markdown file
        with open(md_file, 'r', encoding='utf-8') as file:
            content = file.read()

        # Create a new sheet with the name of the markdown file (without extension)
        sheet_title = os.path.basename(md_file)[:-3] # Remove the last 3 characters (.md)
        sheet = wb.create_sheet(title=sheet_title)

        # Split content into lines and write each line into a new row
        for row_index, line in enumerate(content.splitlines(), start=1):
            # Assuming the content of each line fits within Excel's cell limit
            line = sanitize_for_excel(line)
            sheet.cell(row=row_index, column=1).value = "'"+line

    # Remove the default sheet created by openpyxl if it's untouched
    #if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1 and wb['Sheet'].max_column == 1:
    #    del wb['Sheet']

    # Save the workbook to an xlsx file
    print(save_workbook(wb, filename))

def module_callback(crew,job, crewjob, details):
    """

    """
    console = sys.stdout
    sys.stdout = ComplexLogger(logfile)

    args = argparse.Namespace(
        crew=crew,
        job=job,
        crewjob=crewjob,
        details=details,
    )

    try:
        # Call the function
        gr.Info("Starting process...")
        run_crew(crew,job, crewjob, details)
        print("\nDone.")
        gr.Info("Completed process!")
    except Exception as e:
        gr.Error("Could not complete process!")
        #print(f"ERROR: {e}\n{traceback.format_exc()}")

    sys.stdout = console

def list_xls_files_in_dir(directory):
    """
    This is used to generate the list of template files to choose from
    """
    files_in_directory = os.listdir(directory)
    xlsfiles = [directory + file for file in files_in_directory if file.endswith('.xls') or file.endswith('.xlsx')]
    return xlsfiles

def md_list(items):
        return "\n".join(f"* {item}" for item in items)
    
# Function to get distinct values from a named column in a named sheet
def get_distinct_column_values_by_name(workbook_path, sheet_name, column_name):
    """
    Extracts distinct values from a specified named column in an Excel sheet.

    :param workbook_path: The path to the Excel workbook.
    :param sheet_name: The name of the worksheet to use.
    :param column_name: The name of the column to extract values from.
    :return: A set of distinct values.
    """
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[sheet_name]

    # Find the index of the specified column by name
    column_index = 0
    for col in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
        column_index = column_index +1
        if col[0] == column_name:
            break
    else:
        # If no matching header is found, raise an error
        raise ValueError(f"Column with header '{column_name}' not found.")

    # Use a set to store unique values as sets do not allow duplicates
    distinct_values = set()

    # Iterate over all rows in the specified named column (excluding the header)
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        if cell_value is not None:  # Exclude empty cells
            distinct_values.add(cell_value)

    # Close the workbook after processing
    workbook.close()

    return sorted(list(distinct_values))

def run_crew(crew,job, crewjob, details, input1,input2,input3,input4,input5):
    """
    This is the main function that you will use to run your custom crew.
    """
    (crew, job) = crewjob.split('-', maxsplit=1)   
    crews_dir=f"{CREWS_FOLDER_NAME}.{crew}-{job}"
    select_language='en'

    input_mapping = get_input_mapping(details,input1,input2,input3,input4,input5)
    expanded_details = details.format(**input_mapping)

    # Import the  module dynamically
    crew_module = import_module(f"{crews_dir}.crew")
    CustomCrew = getattr(crew_module, 'CustomCrew')
    custom_crew = CustomCrew(expanded_details, select_language)

    # NOK does not work properly
    #from document_crew import describe_crew_instances2
    #print(describe_crew_instances2(custom_crew))

    # OK, outputs the docstrings with basic config info
    #help(custom_crew)
    #help(custom_crew.agents)
    #help(custom_crew.tasks)

    (result, metrics) = custom_crew.run()

    write_log_sheet(output_log_sheet,details, read_logs(), result['final_output'], json.dumps(metrics, indent=4))
    add_md_files_to_log_sheet(output_log_sheet,f"{CREWS_FOLDER}{crew}-{job}")
 
    return (result['final_output'], metrics)

def get_crew_jobs_list(crewdir):
    crewjobs_list = os.listdir(crewdir)
    crewjobs_list.sort()
    return crewjobs_list

def get_crew_job(crewdir=CREWS_FOLDER):   
    crewjobs_list = get_crew_jobs_list(crewdir)
    crewjob = gr.Dropdown(choices=crewjobs_list , label="Prepared teams")
    return crewjob
    

def setup(template,crew, job):
    """
    This is used to generate a new crew-job combination
    """
    crews_dir = f"{CREWS_FOLDER}{crew}-{job}/"
    create_default_dir(crews_dir)

    read_variables_xls(template,crew, job, crews_dir)
    crewjob = get_crew_job(CREWS_FOLDER)
    
    return("Crew for Job " + crews_dir + " created!" , crewjob)

def get_crews_details(template):
    df = pd.read_excel(template, sheet_name='crewmembers', usecols=['crewmember', 'crew'])

    # Group the DataFrame by 'Crew' and qgents into lists
    grouped_crews = df.groupby('crew')['crewmember'].apply(list)

    # Generate Markdown output
    markdown_output = ""
    for crew, agent in grouped_crews.items():
        # Format each job with its subtasks as a bullet point in Markdown
        markdown_output += f"* {crew} ({', '.join(agent)})\n"

    return markdown_output

def get_jobs_details(template):
    df = pd.read_excel(template, sheet_name='tasks', usecols=['task', 'job'])

    # Group the DataFrame by 'Job' and aggregate subtasks into lists
    grouped_jobs = df.groupby('job')['task'].apply(list)

    # Generate Markdown output
    markdown_output = ""
    for job, task in grouped_jobs.items():
        # Format each job with its subtasks as a bullet point in Markdown
        markdown_output += f"* {job} ({', '.join(task)})\n"

    return markdown_output

def get_crews_jobs_from_template(template, input_crew, input_job):
    """
    This is used to fetch list of available crews and jobs in selected template
    """
    crews_list = get_distinct_column_values_by_name(template, 'crews', 'crew')
    jobs_list = get_distinct_column_values_by_name(template, 'tasks', 'job')
    
    crew = gr.Dropdown(choices=crews_list, label="Select crew")
    job = gr.Dropdown(choices=jobs_list, label="Select job")

    crews_md = gr.Markdown('# CREWS\n' + get_crews_details(template))
    jobs_md = gr.Markdown('# JOBS\n' + get_jobs_details(template))
    
    return (crew, crews_md, job, jobs_md)

def upload_file(in_files):  
    # theoretically allow for multiple, might add output file
    file_paths = [file.name for file in in_files]
    for file in in_files:
        shutil.copy(file, XLS_FOLDER)    
    gr.Info("Files Uploaded!!!")    

    templates_list = list_xls_files_in_dir(XLS_FOLDER)
    template = gr.Dropdown(choices=templates_list, label="1) Select from templates")

    return (file_paths, template)

def extract_variables(details):
    from textwrap import dedent
    
    # Regular expression pattern to find {variable} occurrences
    pattern = re.compile(r'\{(.*?)\}')
    
    # Find all matches of the pattern in the details
    matches = pattern.findall(details)
    
    # Return a list of unique variable names without duplicates
    return sorted(list(set(matches)))

def map_variables_to_ui_fields(description, ui_fields):
    # Extract variables from the description
    variables = extract_variables(description)

    # Map extracted variables to UI fields based on their order
    ui_field_values = {}
    for i, var in enumerate(variables):
        if i < len(ui_fields):
            ui_field_values[i] = var 
        else:
            print(f"Warning: Not enough UI fields to map all variables. Variable '{var}' is ignored.")
            break
    
    return ui_field_values

def get_ui_field_labels():
    # Let's assume these are the fieldnames from your UI fields
    return ['input1', 'input2', 'input3', 'input4', 'input5']

def get_mapped_variables(details):
    # Map variables to UI fields
    return map_variables_to_ui_fields(details,  get_ui_field_labels())

def get_input_mapping(details, input1, input2, input3, input4, input5):
    # actual field names, so we can replace corresponsing labels
    ui_fields =  [ input1, input2, input3, input4, input5]
    variables = extract_variables(details)

    # Map extracted variables to UI fields based on their order
    input_mapping = {}
    for i, var in enumerate(variables):
        if i < len(ui_fields):
            input_mapping[var] = ui_fields[i]
        else:
            print(f"Warning: Not enough UI fields to map all variables. Variable '{var}' is ignored.")
            break
    
    return input_mapping

def parse_details(details):

    # Map variables to UI fields
    mapped_variables = get_mapped_variables(details)
  
    # Zero indexed !!!
    input_vars = len(mapped_variables)

    local_input1 = gr.Textbox(lines=1, label="input x")
    local_input2 = gr.Textbox(lines=1, label="input x")
    local_input3 = gr.Textbox(lines=1, label="input x")
    local_input4 = gr.Textbox(lines=1, label="input x")
    local_input5 = gr.Textbox(lines=1, label="input x")
    if input_vars > 0:
        local_input1 = gr.Textbox(lines=1, visible=True, label=mapped_variables[0])
    if input_vars > 1:
        local_input2 = gr.Textbox(lines=1, visible=True, label=mapped_variables[1])
    if input_vars > 2:
        local_input3 = gr.Textbox(lines=1, visible=True, label=mapped_variables[2])
    if input_vars > 3:
        local_input4 = gr.Textbox(lines=1, visible=True, label=mapped_variables[3])
    if input_vars > 4:
        local_input5 = gr.Textbox(lines=1, visible=True, label=mapped_variables[4])
    
    return (local_input1, local_input2, local_input3, local_input4, local_input5)
    
def get_jobdetails(crewjob):
    def read_prompt_from_disk(file_name):
        with open(file_name, 'r') as file:
            prompt = file.read()
        return prompt

    return gr.Textbox(lines=5, value=read_prompt_from_disk( f"{CREWS_FOLDER}{crewjob}/job_default_prompt.txt"), label="Got default prompt")

global demo

def run_gradio():
    global demo
    with gr.Blocks(theme='freddyaboulton/dracula_revamped', css="#console-logs { background-color: black; }") as demo:
    #with gr.Blocks(theme=gr.themes.Soft(primary_hue="indigo", secondary_hue="slate")) as demo:
        gr.Markdown("# Your CREWAI XLS Runner")
        gr.Markdown("__Easy as 1 - 2 - 3__")
        with gr.Tab("1 - Download xls Template"):
            with gr.Row():
                with gr.Column(scale=1, variant="compact"):            
                    gr.Markdown("### Available Templates")
                    xls_files = list_xls_files_in_dir(XLS_FOLDER)
                    for xls_file in xls_files:
                        gr.File(value=xls_file, label=os.path.basename(xls_file))
        with gr.Tab("2 - Prepare"):
            #with gr.Accordion("Open to Load and generate crews", open=False):
            with gr.Row():
                with gr.Column(scale=1, variant="compact"):            
                    gr.Markdown("### load a new configuration template")
                    xls_template = gr.File()
                    upload_button = gr.UploadButton("Upload xls crewAI template", file_types=["file"], file_count="multiple")
                    gr.Markdown("### Prepare new Crew-Job combination from loaded template")
                    template = gr.Dropdown(choices=templates_list, label="1) Select from templates")
                    upload_button.upload(upload_file, upload_button, outputs=[xls_template, template])
                    read_template_btn = gr.Button("Get Crews and Jobs defined")
                with gr.Column(scale=2, variant="compact"): 
                    #crew = gr.Textbox(label="2) Enter Crew")
                    crews =  gr.Markdown(f"# CREWS {crews_list}")
                    crew = gr.Dropdown(choices=crews_list, label="Select crew")
                with gr.Column(scale=2, variant="compact"): 
                    jobs =  gr.Markdown(f"# JOBS {jobs_list}")
                    #job = gr.Textbox(label="3) Enter Job")
                    job = gr.Dropdown(choices=jobs_list, label="Select job")
                with gr.Column(scale=1, variant="compact"): 
                    setup_btn = gr.Button("Generate Crew-Job combination")
                    setup_result = gr.Markdown(".")


        with gr.Tab("3 - Run"):
            with gr.Row():
                with gr.Column(scale=1, variant="default"):
                    #get_crew_jobs_btn = gr.Button("Get prepared teams and")
                    gr.Markdown("## Complete the prompt ")
                    crewjob = gr.Dropdown(choices=crewjobs_list, label="Select team", allow_custom_value=True)
                    jobdetails = gr.Textbox(lines=5, label="Specify what exactly needs to be done")
                    input1 = gr.Textbox(lines=1, visible=False, label="input 1")
                    input2 = gr.Textbox(lines=1, visible=False, label="input 2")
                    input3 = gr.Textbox(lines=1, visible=False, label="input 3")
                    input4 = gr.Textbox(lines=1, visible=False, label="input 4")
                    input5 = gr.Textbox(lines=1, visible=False, label="input 5")
                    crewjob.change(get_jobdetails, inputs=[crewjob], outputs=[jobdetails])
                    jobdetails.blur(parse_details, inputs=[jobdetails], outputs=[input1,input2,input3,input4,input5])
                with gr.Column(scale=1, variant="default"):
                    gr.Markdown("##  Hit the button")
                    run_crew_btn = gr.Button("Run Crew-Job for job details")
                    metrics = gr.Textbox(lines=2, label="Usage Metrics")
                with gr.Column(scale=2):
                    gr.Markdown("## Wait for results below")
                    gr.Markdown("#### (or watch progress in the console at the bottom)")
                    output = gr.Textbox(lines=20, label="Final output")

            with gr.Row():
                with gr.Column():
                    with gr.Accordion("Console Logs"):
                        # Add logs
                        logs = gr.Code(label="", language="shell", interactive=False, container=True, lines=30, elem_id="console-logs")
                        demo.load(read_logs, None, logs, every=1)
                        demo.load(read_logs, None, logs, every=1)
                        #logs = gr.Textbox()

        read_template_btn.click(get_crews_jobs_from_template, inputs=[template, crew, job], outputs=[crew, crews, job, jobs])
        setup_btn.click(setup, inputs=[template,crew,job], outputs=[setup_result, crewjob])
        #get_crew_jobs_btn.click(get_crew_job, inputs=[], outputs=[crewjob])
        
        #output = run_crew_btn.click(module_callback,[crew,job, crewjob,jobdetails])
        run_crew_btn.click(run_crew,inputs=[crew,job, crewjob,jobdetails,input1,input2,input3,input4,input5], outputs=[output, metrics])

        log = read_logs()

    global demo
    demo.queue().launch(show_error=True)
    #demo.queue().launch(share=True, show_error=True, auth=("user", "pwd"))

#############################################################################
# START PROCESSING
#############################################################################

# Ensure default directories exist
init_default_dirs()

# start logging
logger=init_logging()

crews_list = list()
jobs_list = list()    
templates_list = list_xls_files_in_dir(XLS_FOLDER)
crewjobs_list = get_crew_jobs_list(CREWS_FOLDER)

run_gradio()
